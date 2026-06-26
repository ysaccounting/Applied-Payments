"""
Zone 1 — turn a RAW applied-payments report (CSV) into an enriched review
workbook (.xlsx) with the answer columns + dropdowns + Guidelines tab.

Public API:
    generate_review_workbook(input_path) -> (openpyxl Workbook, data_tab_name)
"""
import os
import csv
import io
import re
import processor as P
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, Color
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

RAW_COLS = ['Company', 'Order#', 'Amount', 'Status', 'TV Fee', 'Payout', 'EventName', 'Performer',
            'Opponent', 'Venue', 'EventDate', 'Section', 'Row', 'Seat', 'Qty', 'CancellationReason',
            'InternalFulfillmentStatus', 'Notes']

# The answer columns that Zone 2 reads back (kept here so both sides agree).
# Order matters — these are written T..X in this exact order.
ANSWER_HEADERS = ['Order Tag', 'Cancelled Out?', 'Already Paid?', 'Cancellation Reason',
                  'Cancelled Old / Paid New?']

# Flag fills (RGB) — one per row category.
FILL_CHARGEBACK = 'FFFFFFCC'   # negative amount / blank company
FILL_NOTFOUND   = 'FFFCE4D6'   # status "Skipped Invoice Not Found"
FILL_MTI        = 'FFDEEBF7'   # status "Skipped Found more than one invoice ..."

# Order-Tag default cancellation reasons (shown in column W when not prefilled).
TAG_DEFAULT_REASON = {
    'Cancelled Event':     'Event Cancelled/Postponed - NA',
    'Problem Order':       'Cancelled by Marketplace - BR',
    'Mutual Cancellation': 'Mutually Cancelled - BR',
}
# Tags whose Already Paid? column should default to Yes.
ALREADY_PAID_YES_TAGS = ('Cancelled Event', 'Mutual Cancellation')


def _realign(fields):
    """Same comma-shift handling the processor uses, returning canonical 18 cols."""
    f = list(fields)
    at = lambda i: f[i] if i < len(f) else ''
    if   P._is_valid_date(at(10)): shift = 0
    elif P._is_valid_date(at(11)): shift = 1
    elif P._is_valid_date(at(12)): shift = 2
    else:                          shift = 0
    eventname = ','.join(f[6:7 + shift])
    tail = f[7 + shift:7 + shift + 11]; tail += [''] * (11 - len(tail))
    return [str(x).strip() for x in (f[0:6] + [eventname] + tail)]


def _to_amt(v):
    try:
        return float(str(v).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip())
    except Exception:
        return None


_DATE_RE = re.compile(r'^\s*(\d{1,2}/\d{1,2}/\d{2,4})')


def _strip_time(s):
    """EventDate display: keep the date, drop any '5:00:00 PM' timestamp."""
    s = str(s).strip()
    if not s:
        return ''
    m = _DATE_RE.match(s)
    return m.group(1) if m else s


def _flag_category(canon):
    """Return 'chargeback' | 'notfound' | 'mti_chargeback' | 'mti' | None for a
    canonical row. Drives both the highlight colour and the sort order.
    A 'more than one invoice' row with a negative amount is a chargeback variant
    ('mti_chargeback', greenish) handled like a regular chargeback; positive
    amounts stay 'mti' (blue)."""
    status = str(canon[3]).strip().lower()
    if 'skipped invoice not found' in status:
        return 'notfound'
    if 'skipped found more than one invoice' in status:
        a = _to_amt(canon[2])
        return 'mti_chargeback' if (a is not None and a < 0) else 'mti'
    a = _to_amt(canon[2])
    if (a is not None and a < 0) or canon[0].strip() == '':
        return 'chargeback'
    return None


# Sort order: chargeback, not-found, MTI-chargeback (negative), MTI (positive), rest.
_SORT_RANK = {'chargeback': 0, 'notfound': 1, 'mti_chargeback': 2, 'mti': 3, None: 4}
# 'mti_chargeback' uses accent3 (9BBB59) lightened — a greenish highlight.
FILL_MTI_CHARGEBACK = Color(theme=6, tint=0.6)
_FLAG_FILL = {'chargeback': FILL_CHARGEBACK, 'notfound': FILL_NOTFOUND,
              'mti_chargeback': FILL_MTI_CHARGEBACK, 'mti': FILL_MTI}


def _read_rows(input_path):
    """Read the raw report resiliently (handles stray CP1252 bytes / BOM)."""
    with open(input_path, 'rb') as fh:
        data = fh.read()
    if data[:3] == b'\xef\xbb\xbf':
        data = data[3:]
    text = data.decode('utf-8', errors='utf8_cp1252')
    return list(csv.reader(io.StringIO(text)))


def _tab_name(input_path):
    try:
        net = P.parse_filename(os.path.basename(input_path))[0]
        name = re.sub(r'[^A-Za-z0-9 ]', '', str(net)).strip()[:31]
        return name or 'Report'
    except Exception:
        return 'Report'


# Order Tags offered on every network, plus tags that only occur on one network's
# reports (per the Guidelines tab — StubHub Loan, TradeDesk Fees, Due from/to
# TickPick only ever show up on StubHub / TicketsNow / TickPick reports).
ORDER_TAGS_BASE = ['Cancelled Event', 'Discount', 'More Than One Invoice', 'Mutual Cancellation',
                   'Not Found', 'Problem Order']
NETWORK_TAGS = {              # substring of the normalised network key -> extra tag
    'stubhub':    'StubHub Loan',
    'tickpick':   'Due from/to TickPick',
    'ticketsnow': 'TradeDesk Fees',
}


def _net_key(input_path):
    """Normalised network key for the file (lowercase, no spaces/parens)."""
    try:
        net = str(P.parse_filename(os.path.basename(input_path))[0])
    except Exception:
        net = ''
    return net.lower().replace('(', '').replace(')', '').replace(' ', '')


def _order_tags(input_path):
    """Order-Tag dropdown list for this file's network: the base tags plus any
    network-specific tag, kept alphabetical to match the sample."""
    key = _net_key(input_path)
    tags = list(ORDER_TAGS_BASE)
    for sub, tag in NETWORK_TAGS.items():
        if sub in key:
            tags.append(tag)
    return sorted(tags)


def _order_is_zero(v):
    """True when the Order# is missing or zero."""
    s = str(v).strip()
    if s == '':
        return True
    try:
        return float(s) == 0
    except ValueError:
        return False


def _norm_reason(s):
    """Collapse whitespace + lowercase for tolerant reason matching."""
    return ''.join(str(s).split()).lower()


def _default_order_tag(cat, canon, net_key):
    """Default Order Tag (col T) prefill for a flagged row. The reviewer can still
    override via the dropdown.
      • More-than-one-invoice rows -> More Than One Invoice
      • Not-Found rows             -> Not Found, EXCEPT a TicketsNow Not-Found with
                                      no order # (genuinely ambiguous — could be Not
                                      Found or TradeDesk Fees) -> left blank
      • Chargebacks                -> by preloaded cancellation reason:
            Mutually Cancelled - BR        -> Mutual Cancellation
            Event Cancelled/Postponed - NA -> Cancelled Event
            any other non-blank reason     -> Problem Order
            blank reason                   -> no default (left for the reviewer)
    """
    if cat == 'mti':
        return 'More Than One Invoice'
    if cat == 'notfound':
        if 'ticketsnow' in net_key and _order_is_zero(canon[1]):
            return ''
        return 'Not Found'
    if cat in ('chargeback', 'mti_chargeback'):
        reason = _norm_reason(canon[15])
        if not reason:
            return ''
        if reason == _norm_reason(TAG_DEFAULT_REASON['Mutual Cancellation']):
            return 'Mutual Cancellation'
        if reason == _norm_reason(TAG_DEFAULT_REASON['Cancelled Event']):
            return 'Cancelled Event'
        return 'Problem Order'
    return ''


def generate_review_workbook(input_path):
    """Build and return (Workbook, data_tab_name) for the given raw report CSV."""
    all_rows = _read_rows(input_path)
    rows = [_realign(r) for r in all_rows[1:]]
    # Flagged rows (chargeback, then not-found, then more-than-one-invoice) to the
    # top, each group sorted by amount ascending; unflagged rows last.
    rows.sort(key=lambda c: (_SORT_RANK[_flag_category(c)],
                             _to_amt(c[2]) if _to_amt(c[2]) is not None else 0))
    net_key = _net_key(input_path)

    wb = Workbook(); ws = wb.active; ws.title = _tab_name(input_path)
    FONT = Font(name='Arial', size=10)
    HFONT = Font(name='Arial', size=10, bold=True)
    HEAD_FILL = PatternFill('solid', fgColor='FFD9E1F2')
    GRAY = PatternFill('solid', fgColor='FFD9D9D9')
    thin = Side(style='thin', color='BFBFBF'); border = Border(thin, thin, thin, thin)

    # A-R raw | S sep | T Order Tag | U Cancelled Out? | V Already Paid? |
    # W Cancellation Reason | X Cancelled Old / Paid New?
    headers = RAW_COLS + [''] + ANSWER_HEADERS
    for ci, name in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=name); c.font = HFONT
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.fill = GRAY if ci == 19 else HEAD_FILL
        c.border = border

    UNLOCKED = Protection(locked=False)
    LOCKED = Protection(locked=True)
    # Column indexes for the answer block.
    T, U, V, W, X = 20, 21, 22, 23, 24
    for ri, canon in enumerate(rows, start=2):
        cat = _flag_category(canon)
        flag = cat is not None
        flag_fill = PatternFill('solid', fgColor=_FLAG_FILL[cat]) if flag else None
        for ci in range(1, 19):                       # A-R
            val = canon[ci - 1]
            if ci in (3, 5, 6):                        # Amount, TV Fee, Payout -> numeric, 2 dp
                a = _to_amt(val); val = a if a is not None else val
            if ci == 11:
                val = _strip_time(val)
            cell = ws.cell(row=ri, column=ci, value=val); cell.font = FONT
            if ci in (3, 5, 6):
                cell.number_format = '#,##0.00'
            if ci == 3:
                cell.alignment = Alignment(horizontal='center')
            if flag:
                cell.fill = flag_fill
            cell.protection = UNLOCKED
        sep = ws.cell(row=ri, column=19, value=' '); sep.fill = GRAY; sep.protection = UNLOCKED

        prefill_reason = canon[15].strip()            # raw CancellationReason (col P)
        for ci in (T, U, V, W, X):
            cell = ws.cell(row=ri, column=ci); cell.font = FONT
            if flag:
                cell.fill = flag_fill
            # W (Cancellation Reason) stays editable on every row; the other
            # answer cells are only editable on flagged rows.
            cell.protection = UNLOCKED if (flag or ci == W) else LOCKED

        if flag:
            # Default the Order Tag (the reviewer can still change it).
            default_tag = _default_order_tag(cat, canon, net_key)
            if default_tag:
                ws.cell(row=ri, column=T, value=default_tag)
            # Cancelled Out? + Cancellation Reason prefill from TicketVault.
            if prefill_reason:
                ws.cell(row=ri, column=U, value='Yes')
                ws.cell(row=ri, column=W, value=prefill_reason)
            else:
                # No TV reason yet — show a tag-based default once a tag is chosen.
                ws.cell(row=ri, column=W,
                        value=(f'=IF($T{ri}="Cancelled Event","{TAG_DEFAULT_REASON["Cancelled Event"]}",'
                               f'IF($T{ri}="Problem Order","{TAG_DEFAULT_REASON["Problem Order"]}",'
                               f'IF($T{ri}="Mutual Cancellation","{TAG_DEFAULT_REASON["Mutual Cancellation"]}","")))'))
            # Already Paid? defaults to Yes for cancelled events and mutual cancellations.
            paid_cond = ','.join(f'$T{ri}="{t}"' for t in ALREADY_PAID_YES_TAGS)
            ws.cell(row=ri, column=V, value=f'=IF(OR({paid_cond}),"Yes","")')

    last = len(rows) + 1

    # ── Dropdowns ─────────────────────────────────────────────────────────────
    def dv(formula):
        d = DataValidation(type='list', formula1=formula, allow_blank=True,
                           showErrorMessage=True, errorStyle='stop')
        d.error = 'Pick a value from the list.'; d.errorTitle = 'Invalid entry'
        ws.add_data_validation(d); return d
    dv('"' + ','.join(_order_tags(input_path)) + '"').add(f'T2:T{last}')
    yes_dv = dv('"Yes"'); yes_dv.add(f'U2:U{last}'); yes_dv.add(f'X2:X{last}')
    dv('"Yes,No"').add(f'V2:V{last}')

    widths = {'A': 16, 'B': 14, 'C': 12, 'D': 74, 'E': 8, 'F': 10, 'G': 32, 'H': 20,
              'J': 22, 'K': 18, 'L': 8, 'M': 6, 'N': 8, 'O': 5, 'P': 22, 'Q': 20,
              'R': 30, 'S': 3, 'T': 22, 'U': 16, 'V': 15, 'W': 34, 'X': 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions['G'].hidden = True   # EventName hidden by default

    ws.freeze_panes = 'I2'   # pin columns A-H while scrolling right to the answers
    ws.auto_filter.ref = 'A1:R1'
    ws.protection = SheetProtection(sheet=True,
        selectLockedCells=False, selectUnlockedCells=False,
        sort=False, autoFilter=False,
        formatCells=False, formatColumns=False, formatRows=False)

    # ── Blacked-out cells (nothing to do there for the chosen tag) ────────────
    black = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    blackfont = Font(name='Arial', size=10, color='FF000000')
    # U,V,W go black for tags that need no per-column input.
    ws.conditional_formatting.add(f'U2:W{last}',
        FormulaRule(formula=['OR($T2="Discount",$T2="Not Found",$T2="More Than One Invoice",'
                             '$T2="Due from/to TickPick",$T2="StubHub Loan",$T2="TradeDesk Fees")'],
                    fill=black, font=blackfont, stopIfTrue=False))
    # X is only used for More Than One Invoice — black for any other selected tag.
    ws.conditional_formatting.add(f'X2:X{last}',
        FormulaRule(formula=['AND($T2<>"",$T2<>"More Than One Invoice")'],
                    fill=black, font=blackfont, stopIfTrue=False))

    _build_guidelines_tab(wb)
    return wb, ws.title


# ── Guidelines tab ────────────────────────────────────────────────────────────
_C_TITLE = 'FF1F4E78'   # dark blue tag titles
_YEL, _PEACH, _BLUE = 'FFFFFFCC', 'FFFCE4D6', 'FFDEEBF7'


def _if(b=False, color=None):
    return InlineFont(rFont='Arial', sz=10, b=b, color=color)


def _build_guidelines_tab(wb):
    rs = wb.create_sheet('Guidelines')
    FONT = Font(name='Arial', size=10)
    TITLE = Font(name='Arial', size=13, bold=True)
    wrapL = Alignment(wrap_text=True, vertical='center', horizontal='left')
    wrapC = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrapTop = Alignment(wrap_text=True, vertical='top')

    def fill(hexrgb):
        return PatternFill('solid', fgColor=hexrgb)

    # Title / intro — only the word "Guidelines" is a big bold title; the rest is
    # a normal instruction line.
    rs['A1'] = CellRichText([
        TextBlock(InlineFont(rFont='Arial', sz=13, b=True), 'Guidelines'),
        TextBlock(_if(), ' - complete columns T thru X on the first tab per the guidelines below. '
                         'Some columns come prefilled based on certain info in the raw applied payments '
                         'report from TicketVault.'),
    ])
    rs['A1'].font = FONT; rs.row_dimensions[1].height = 16.5

    # Colour legend
    rs['A2'] = 'Chargeback orders (negative amounts)'
    rs['A2'].font = FONT; rs['A2'].fill = fill(_YEL)
    rs['A3'] = 'Not Found orders (including monthly Due from/to TickPick adjustment and TradeDesk Fees)'
    rs['A3'].font = FONT; rs['A3'].fill = fill(_PEACH)
    rs['A4'] = 'Skipped More Than One Invoice Found orders and chargeback.'
    rs['A4'].font = FONT; rs['A4'].fill = fill(FILL_MTI_CHARGEBACK)
    rs['A5'] = 'Skipped More Than One Invoice Found orders. Column X is only used for these orders'
    rs['A5'].font = FONT; rs['A5'].fill = fill(_BLUE)

    # Note on split chargebacks
    rs['A7'] = ('In cases where a chargeback is split into two rows - one a payout recoup and one a '
                'cancellation fee - use Problem Order tag on both rows. For the recoup row, put Yes for '
                'already paid. For the cancellation fee row, put No for already paid.')
    rs['A7'].font = FONT

    # Detail table header
    RHEAD = PatternFill('solid', fgColor='FF4472C4')
    RH = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    a9 = rs['A9']; a9.value = 'Order Tags'; a9.font = RH; a9.fill = RHEAD
    a9.alignment = Alignment(horizontal='center', vertical='center')
    b9 = rs['B9']; b9.value = 'What Zone 2 does to that row'; b9.font = RH; b9.fill = RHEAD
    b9.alignment = Alignment(horizontal='left', vertical='center')

    def title_cell(title, note):
        return CellRichText([
            TextBlock(_if(b=True, color=_C_TITLE), title + '\n'),
            TextBlock(_if(), note),
        ])

    # (row, fill, A rich title/note, B body[str or CellRichText], height)
    rows = [
        (10, _YEL,
         title_cell('Cancelled Event', '(the full chargeback amount is a payout recoup with no cancellation fee)'),
         '>Nothing changes on the row.\n'
         '>If the order was not already cancelled out from TicketVault, then cancel it and put Yes in Cancelled Out? Column.\n'
         '>Use "Event Cancelled/Postponed - NA" for the cancellation reason in TicketVault.',
         44.25),
        (11, _YEL,
         title_cell('Discount', '(the full chargeback amount is a cancellation fee — other three columns blacked out)'),
         CellRichText([
             '>"-Fee" is added to the Company name so that it\'s treated as a cancellation fee. ',
             TextBlock(_if(b=True), 'DO NOT'),
             TextBlock(_if(), " cancel out the order from TicketVault.\n"
                              ">Discounts are to compensate the buyers for wrong tickets delivered so that they "
                              "won't cancel the sale. Includes shipping fees too.\n"
                              '  The networks might notate these as "discount" or "alternative tickets" etc.\n'
                              ">Amounts should be relatively small compared to the sales price."),
         ]),
         78.75),
        (12, _YEL,
         title_cell('Mutual Cancellation', '(the full chargeback amount is a payout recoup with no cancellation fee)'),
         '>Nothing changes on the row.\n'
         '>If the order was not already cancelled out from TicketVault, then cancel it and put Yes in Cancelled Out? Column.\n'
         '>Use "Mutually Cancelled - BR" for the cancellation reason in TicketVault.',
         44.25),
        (13, _YEL,
         title_cell('Problem Order, and Already Paid? = No', '(the full chargeback amount is a cancellation fee)'),
         '>"-Fee" is added to the Company name so that it\'s treated as a cancellation fee.\n'
         '>If the order was not already cancelled out from TicketVault, then cancel it and put Yes in Cancelled Out? column.\n'
         '>Use "Cancelled by Marketplace - BR" for the cancellation reason in TicketVault.',
         56.25),
        (14, _YEL,
         title_cell('Problem Order, and Already Paid? = Yes', '(the chargeback amount is a payout recoup + cancellation fee)'),
         '>The row is split into two negative lines:\n'
         '    Line 1 = the Payout amount, as a negative, with the original Company.\n'
         '    Line 2 = the remainder, under the Company with "-Fee" added so that it\'s treated as a cancellation fee.\n'
         '    These two lines add back to the original Amount.\n'
         '>If the order was not already cancelled out from TicketVault, then cancel it and put Yes in Cancelled Out? column.\n'
         '>Use "Cancelled by Marketplace - BR" for the cancellation reason in TicketVault.',
         76.5),
        (15, _YEL,
         title_cell('StubHub Loan', '(daily loan repayments from Y&S to StubHub)'),
         '>Replaces the Company and only displays on StubHub reports. No further action required.\n'
         '>Repayments are assigned to random amounts and order numbers (sometimes for tens of thousands of dollars per order) and will not\n'
         '  show signs of being problem orders when searched in Gmail.',
         51.0),
        (16, _YEL,
         title_cell('TradeDesk Fees', '(monthly fulfillment fees paid from Y&S to TicketsNow)'),
         '>Replaces the Company with Other Fees and only displays on TicketsNow reports. No further action required.\n'
         '>Shows up as Not Found chargeback with no order #, and usually for tens of thousands of dollars.',
         25.5),
        (17, _PEACH,
         title_cell('Due from/to TickPick', '(monthly repayment to Y&S for theft loss)'),
         '>Replaces the Company and only displays on TickPick reports. No further action required.\n'
         '>Shows up as Not Found payment, but hopefully with order # DueFromTickPick.',
         30.0),
        (18, _PEACH,
         title_cell('Not Found', '(orders with status "Skipped Invoice Not Found")'),
         '>Replaces the Company. No further action required.',
         30.0),
        (19, FILL_MTI_CHARGEBACK,
         title_cell('More Than One Invoice and Chargeback',
                    '(orders with status "Skipped Found more than one invoice with this Ext Order Number and Client"\n'
                    'and negative amount)'),
         '>Treat the same as regular chargebacks for cancelled events and problem orders.',
         42.0),
        (20, _BLUE,
         title_cell('More Than One Invoice',
                    '(orders with status "Skipped Found more than one invoice with this Ext Order Number and Client")'),
         '>Cancel the older invoice in TicketVault and mark paid the newer invoice in TicketVault, then put Yes in Column X.\n'
         '>Use "Adjusting PO Details - Will be Re-Invoiced - BR" for the cancellation reason in TicketVault.',
         39.75),
    ]
    for r, fl, a_val, b_val, h in rows:
        a = rs.cell(row=r, column=1, value=a_val); a.fill = fill(fl); a.alignment = wrapC; a.font = FONT
        b = rs.cell(row=r, column=2, value=b_val); b.fill = fill(fl); b.alignment = wrapL; b.font = FONT
        rs.row_dimensions[r].height = h

    # Blocking conditions
    rs['A22'] = 'What blocks Zone 2 processing'
    rs['A22'].font = Font(name='Arial', size=11, bold=True, color='FF9C0006')
    blocks = [
        'Any highlighted row without an Order Tag assigned',
        'A Problem Order row without an answer for Already Paid?',
        "A Problem Order or Cancelled Event or Mutual Cancellation row that doesn't say Yes for Cancelled Out?",
        "A Problem Order or Cancelled Event or Mutual Cancellation row that doesn't have a cancellation reason filled in",
        "A More Than One Invoice row that doesn't say Yes for Cancelled Old / Paid New?",
    ]
    r = 23
    for b in blocks:
        c = rs.cell(row=r, column=1, value='\u2022  ' + b)
        c.font = FONT; c.alignment = wrapTop
        rs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
    rs.merge_cells(start_row=28, start_column=1, end_row=28, end_column=2)

    rs.column_dimensions['A'].width = 83.28515625
    rs.column_dimensions['B'].width = 113.0
    return rs
