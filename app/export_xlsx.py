"""Excel export for the review grids.

Two sheets every time:

  1. the rows currently on the grid -- ALL of them, not the page you can see,
     because "export what I'm looking at" means the filtered set, and the page
     size is a display limit rather than part of the question;
  2. the filters that produced it, plus when it was taken.

The second sheet exists because a spreadsheet with no provenance is the kind of
thing that gets emailed on and argued about. "374 transactions" means nothing
without knowing it was Divvy only, refunds only, for one week.
"""
from __future__ import annotations

from datetime import datetime, timezone, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Chicago, without pulling in a tz database: CST is UTC-6, CDT UTC-5, and US
# DST runs from the second Sunday in March to the first Sunday in November.
_CENTRAL_STD = timedelta(hours=-6)
_CENTRAL_DST = timedelta(hours=-5)


def _nth_weekday(year: int, month: int, weekday: int, n: int) -> datetime:
    d = datetime(year, month, 1)
    shift = (weekday - d.weekday()) % 7
    return d + timedelta(days=shift + 7 * (n - 1))


def to_central(dt: datetime | None) -> datetime | None:
    """A UTC timestamp as America/Chicago wall time."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    y = dt.year
    start = _nth_weekday(y, 3, 6, 2) + timedelta(hours=2)    # 2am 2nd Sun Mar
    end = _nth_weekday(y, 11, 6, 1) + timedelta(hours=2)     # 2am 1st Sun Nov
    offset = _CENTRAL_DST if start <= dt < end else _CENTRAL_STD
    return dt + offset


def fmt_central(dt: datetime | None) -> str:
    local = to_central(dt)
    return local.strftime("%m/%d/%Y %I:%M %p") if local else ""


_HEAD_FILL = PatternFill("solid", fgColor="1F4E5F")
_HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_BODY_FONT = Font(name="Arial", size=10)
_LABEL_FONT = Font(name="Arial", bold=True, size=10)


# Excel number formats. Dates are written as real dates and amounts as real
# numbers, so the spreadsheet can sort and total them -- formatting them as
# text would look right and behave wrong.
DATE_FMT = "m/d/yyyy"
MONEY_FMT = '$#,##0.00;($#,##0.00)'
STAMP_FMT = "m/d/yyyy h:mm AM/PM"


def _write_sheet(ws, headers: list[str], rows: list[list],
                 date_cols=(), money_cols=(), stamp_cols=()) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEAD_FILL
        cell.font = _HEAD_FONT
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    for idx in date_cols:
        for cell in ws[get_column_letter(idx + 1)][1:]:
            cell.number_format = DATE_FMT
    for idx in stamp_cols:
        for cell in ws[get_column_letter(idx + 1)][1:]:
            cell.number_format = STAMP_FMT
    for idx in money_cols:
        for cell in ws[get_column_letter(idx + 1)][1:]:
            cell.number_format = MONEY_FMT
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        widest = len(str(headers[col - 1]))
        for r in rows:
            if col - 1 < len(r):
                widest = max(widest, len(str(r[col - 1] if r[col - 1] is not None else "")))
        # A cap, or one long memo makes a column nobody can see past.
        ws.column_dimensions[letter].width = min(max(widest + 2, 10), 55)
        for cell in ws[letter][1:]:
            cell.font = _BODY_FONT
    ws.freeze_panes = "A2"


def build_workbook(title: str, headers: list[str], rows: list[list],
                   filters: list[tuple[str, str]],
                   date_cols=(), money_cols=(), stamp_cols=()) -> bytes:
    """Rows on one sheet, the filters that produced them on another."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Export"
    _write_sheet(ws, headers, rows, date_cols, money_cols, stamp_cols)

    meta = wb.create_sheet("Filters")
    meta["A1"] = "Export details"
    meta["A1"].font = Font(name="Arial", bold=True, size=12)
    line = 3
    # UTC on the sheet: the file gets emailed around, and a zone-less local
    # time is ambiguous the moment it leaves the machine that made it.
    pairs = [("Exported", datetime.utcnow().strftime("%m/%d/%Y %I:%M %p") + " UTC"),
             ("Rows exported", str(len(rows)))] + list(filters)
    for label, value in pairs:
        meta.cell(row=line, column=1, value=label).font = _LABEL_FONT
        c = meta.cell(row=line, column=2, value=value)
        c.font = _BODY_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        line += 1
    meta.column_dimensions["A"].width = 26
    meta.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(prefix: str, company: str) -> str:
    stamp = to_central(datetime.utcnow()).strftime("%Y%m%d-%H%M")
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in company)[:40]
    return f"{prefix}_{safe}_{stamp}.xlsx"
