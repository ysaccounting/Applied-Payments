"""
Zone 2 intake — read a FILLED Zone-1 review workbook (.xlsx), apply the
Order-Tag rules to the highlighted (flagged) rows, and return a DataFrame in
the exact schema the existing processor.process() already expects.

Answer columns (Zone-1 layout, fixed positions):
    T Order Tag | U Cancelled Out? | V Already Paid? | W Cancellation Reason |
    X Cancelled Old / Paid New?

Order-Tag actions applied to FLAGGED rows (see the Guidelines tab):
    Cancelled Event ............ no change (payout recoup, no fee)
    Discount ................... Company := Company + "-Fee"
    Problem Order, Paid? = No .. Company := Company + "-Fee"
    Problem Order, Paid? = Yes . SPLIT into two negative lines:
                                   line1 = -|Payout|, original Company
                                   line2 = Amount + |Payout|, Company + "-Fee"
    StubHub Loan ............... Company := "StubHub Loan"
    TradeDesk Fees ............. Company := "Other Fees"
    Due from/to TickPick ....... Company := "Due from/to TickPick"
    Not Found .................. Company := "Not Found"
    More Than One Invoice ...... pass through unchanged (needs Col X = Yes)

Non-flagged rows pass through untouched; any answers on them are ignored.
Cancellation Reason (col W) becomes the processor's Reason field.

Blocks (raise ValueError) — matching the "What blocks Zone 2 processing"
list at the bottom of the Guidelines tab:
    1. Any highlighted row without an Order Tag assigned
    2. A Problem Order row with no answer for Already Paid?
    3. A Problem Order or Cancelled Event row that doesn't say Yes for Cancelled Out?
    4. A Problem Order or Cancelled Event row that doesn't have a cancellation reason
    5. A More Than One Invoice row that doesn't say Yes for Cancelled Old / Paid New?
"""
import os
import re
import pandas as pd
from openpyxl import load_workbook

from zone1 import RAW_COLS, ANSWER_HEADERS   # keep both sides in agreement

PROC_COLS = RAW_COLS + ['Reason']            # what processor.process() expects (Reason appended)

# Order Tags that simply replace the Company name (no other action / no fee math).
COMPANY_REPLACEMENT = {
    'StubHub Loan':         'StubHub Loan',
    'TradeDesk Fees':       'Other Fees',
    'Due from/to TickPick': 'Due from/to TickPick',
    'Not Found':            'Not Found',
}


def _to_amt(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip())
    except Exception:
        return None


def _s(v):
    return '' if v is None else str(v).strip()


def _data_sheet(wb):
    for s in wb.worksheets:
        if s.title not in ('Guidelines', 'Rules'):
            return s
    return wb.worksheets[0]


def _flagged(company, amount, status):
    """Mirror Zone 1's highlight logic: chargeback / Not Found / More Than One Invoice."""
    st = status.lower()
    if 'skipped invoice not found' in st:
        return True
    if 'skipped found more than one invoice' in st:
        return True
    return (amount is not None and amount < 0) or company == ''


def _locate_columns(ws):
    """1-based column indexes for raw + answer fields, by header name, falling back
    to the fixed Zone-1 layout (A-R raw, T-X answers)."""
    header = {}
    for c in range(1, ws.max_column + 1):
        name = _s(ws.cell(row=1, column=c).value)
        if name:
            header.setdefault(name, c)
    cols = {}
    for i, name in enumerate(RAW_COLS, start=1):
        cols[name] = header.get(name, i)
    fixed = {'Order Tag': 20, 'Cancelled Out?': 21, 'Already Paid?': 22,
             'Cancellation Reason': 23, 'Cancelled Old / Paid New?': 24}
    for name, pos in fixed.items():
        cols[name] = header.get(name, pos)
    return cols


def read_filled_zone1(xlsx_path, original_filename):
    """
    -> (raw_df, processor_filename, zone1_values)
       raw_df            : DataFrame in PROC_COLS schema, rules applied.
       processor_filename: filename handed to process() (the _zone1 suffix removed).
       zone1_values      : [header_row, *data_rows] (A-R + T-X) as submitted, for the tab.
    Raises ValueError on the block conditions above.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = _data_sheet(wb)
    col = _locate_columns(ws)

    def cell(r, name):
        return ws.cell(row=r, column=col[name]).value

    out_rows = []
    snapshot = [list(RAW_COLS) + list(ANSWER_HEADERS)]

    missing_tag, paid_err, cancelled_err, reason_err, mti_err = [], [], [], [], []

    for r in range(2, ws.max_row + 1):
        company = _s(cell(r, 'Company'))
        amount = _to_amt(cell(r, 'Amount'))
        order = _s(cell(r, 'Order#'))
        status = _s(cell(r, 'Status'))
        if amount is None and company == '' and order == '':
            continue  # truly empty trailing row

        tag = _s(cell(r, 'Order Tag'))
        cancelled = _s(cell(r, 'Cancelled Out?'))
        paid = _s(cell(r, 'Already Paid?'))
        reason = _s(cell(r, 'Cancellation Reason'))
        old_new = _s(cell(r, 'Cancelled Old / Paid New?'))

        snapshot.append([_s(cell(r, n)) if n != 'Amount' else (amount if amount is not None else _s(cell(r, n)))
                         for n in RAW_COLS] + [tag, cancelled, paid, reason, old_new])

        base = {n: _s(cell(r, n)) for n in RAW_COLS}
        base['Amount'] = amount if amount is not None else _s(cell(r, 'Amount'))
        base['Reason'] = reason

        if not _flagged(company, amount, status):
            out_rows.append(base)             # ignore any answers on non-flagged rows
            continue

        where = f"row {r} (Order# {order or 'blank'}, Amount {amount})"

        # ── the five block conditions from the Guidelines tab ─────────────────
        if tag == '':
            missing_tag.append(where)
            continue                          # nothing else to validate without a tag
        if tag == 'Problem Order' and paid == '':
            paid_err.append(f"row {r} (Order# {order or 'blank'})")
        if tag in ('Problem Order', 'Cancelled Event') and cancelled != 'Yes':
            cancelled_err.append(f"row {r} (Order# {order or 'blank'}, {tag})")
        if tag in ('Problem Order', 'Cancelled Event') and reason == '':
            reason_err.append(f"row {r} (Order# {order or 'blank'}, {tag})")
        if tag == 'More Than One Invoice' and old_new != 'Yes':
            mti_err.append(f"row {r} (Order# {order or 'blank'})")

        # ── transformation by Order Tag ──────────────────────────────────────
        if tag == 'Cancelled Event':
            out_rows.append(base)
        elif tag == 'Discount':
            base['Company'] = company + '-Fee'
            out_rows.append(base)
        elif tag == 'Problem Order':
            if paid == 'Yes':
                payout = abs(_to_amt(cell(r, 'Payout')) or 0.0)
                line1 = dict(base); line1['Amount'] = -payout
                line2 = dict(base); line2['Amount'] = (amount or 0.0) + payout
                line2['Company'] = company + '-Fee'
                out_rows.append(line1)
                out_rows.append(line2)
            elif paid == 'No':
                base['Company'] = company + '-Fee'
                out_rows.append(base)
            # paid blank already recorded in paid_err above
        elif tag in COMPANY_REPLACEMENT:
            base['Company'] = COMPANY_REPLACEMENT[tag]
            out_rows.append(base)
        elif tag == 'More Than One Invoice':
            out_rows.append(base)             # pass through; needs Col X = Yes
        else:
            out_rows.append(base)             # unknown tag — leave row unchanged

    problems = []
    if missing_tag:
        problems.append("These highlighted rows have no Order Tag assigned. Pick a tag in column T:\n  • "
                        + "\n  • ".join(missing_tag))
    if paid_err:
        problems.append("These Problem Order rows have no answer for Already Paid?. Fill column V with Yes or No:\n  • "
                        + "\n  • ".join(paid_err))
    if cancelled_err:
        problems.append("These Problem Order / Cancelled Event rows don't say Yes for Cancelled Out?. "
                        "Confirm the order is cancelled out in TicketVault, then set column U to Yes:\n  • "
                        + "\n  • ".join(cancelled_err))
    if reason_err:
        problems.append("These Problem Order / Cancelled Event rows have a blank Cancellation Reason. "
                        "Fill column W:\n  • "
                        + "\n  • ".join(reason_err))
    if mti_err:
        problems.append("These More Than One Invoice rows don't say Yes for Cancelled Old / Paid New?. "
                        "Cancel the older invoice and mark the newer one paid in TicketVault, then set column X to Yes:\n  • "
                        + "\n  • ".join(mti_err))
    if problems:
        raise ValueError("Can't process —\n\n" + "\n\n".join(problems))

    raw_df = pd.DataFrame(out_rows, columns=PROC_COLS)

    fn = re.sub(r'_zone1', '', original_filename, flags=re.I)
    processor_filename = os.path.splitext(fn)[0] + '.csv'
    return raw_df, processor_filename, snapshot


def looks_like_zone1(xlsx_path):
    """True if the workbook carries the Zone-1 answer columns."""
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = _data_sheet(wb)
        hdr = {_s(ws.cell(row=1, column=c).value) for c in range(1, (ws.max_column or 0) + 1)}
        wb.close()
        return 'Order Tag' in hdr and 'Cancellation Reason' in hdr
    except Exception:
        return False


def add_zone1_output_tab(wb, snapshot, title='Zone 1 Output'):
    """Append the as-submitted Zone-1 data (header + rows) as a tab in an output workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment
    ws = wb.create_sheet(title)
    HFONT = Font(name='Arial', size=10, bold=True)
    FONT = Font(name='Arial', size=10)
    HEAD = PatternFill('solid', fgColor='FFD9E1F2')
    FILLS = {'chargeback': PatternFill('solid', fgColor='FFFFFFCC'),
             'notfound':   PatternFill('solid', fgColor='FFFCE4D6'),
             'mti':        PatternFill('solid', fgColor='FFDEEBF7')}

    def _cat(company, amount, status):
        st = status.lower()
        if 'skipped invoice not found' in st:
            return 'notfound'
        if 'skipped found more than one invoice' in st:
            return 'mti'
        if (amount is not None and amount < 0) or company == '':
            return 'chargeback'
        return None

    header = snapshot[0]
    for ci, name in enumerate(header, 1):
        c = ws.cell(row=1, column=ci, value=name); c.font = HFONT; c.fill = HEAD
        c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, row in enumerate(snapshot[1:], start=2):
        amt = row[2] if isinstance(row[2], (int, float)) else _to_amt(row[2])
        cat = _cat(_s(row[0]), amt, _s(row[3]))
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val); c.font = FONT
            if ci == 3 and isinstance(val, (int, float)):
                c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='center')
            if cat:
                c.fill = FILLS[cat]
    widths = {'A': 16, 'B': 14, 'C': 12, 'D': 14, 'E': 8, 'F': 10, 'G': 32, 'H': 20, 'J': 22,
              'K': 18, 'L': 8, 'M': 6, 'N': 8, 'O': 5, 'P': 22, 'Q': 20, 'R': 30,
              'S': 22, 'T': 16, 'U': 15, 'V': 34, 'W': 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'P2'
    return ws
