import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
import os
import io
import codecs


# TicketVault/EvoPay files are usually UTF-8, but Excel sometimes injects stray
# Windows-1252 "smart" punctuation bytes (e.g. 0x92 = a curly apostrophe) that
# break a strict UTF-8 read. This error handler decodes ONLY the stray bytes as
# CP1252, so genuine UTF-8 (accented venue/performer names like "Rüf") is kept
# intact while the Windows punctuation is recovered instead of crashing the upload.
def _utf8_cp1252_fallback(err):
    return (err.object[err.start:err.end].decode("cp1252", errors="replace"), err.end)


codecs.register_error("utf8_cp1252", _utf8_cp1252_fallback)


def _read_csv_text(path):
    """Read a CSV file's text resiliently (UTF-8 with Windows-1252 recovery)."""
    with open(path, "rb") as fh:
        raw = fh.read()
    if raw.startswith(b"\xef\xbb\xbf"):  # strip UTF-8 BOM if present
        raw = raw[3:]
    return raw.decode("utf-8", errors="utf8_cp1252")


def _read_csv_resilient(path, **kwargs):
    """pd.read_csv that tolerates stray Windows-1252 bytes in an otherwise UTF-8 file."""
    return pd.read_csv(io.StringIO(_read_csv_text(path)), **kwargs)


def _to_amount(v):
    """Parse a money value to float. Handles plain numbers and strings that
    include thousands commas, a leading $, surrounding spaces, or accounting
    parentheses for negatives (e.g. '-1,234.50', '$1,234', '(123.45)'). Blank
    or unparseable values become 0.0. Some networks (e.g. GoTickets) export the
    Amount column with thousands commas, which pandas leaves as text."""
    if isinstance(v, (int, float)):
        return float(v) if pd.notna(v) else 0.0
    s = str(v).strip()
    if s == "" or s.lower() == "nan":
        return 0.0
    negative = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if negative else val


# ── Company mapping (static) ──────────────────────────────────────────────────
COMPANY_MAPPING = {
    "Damon and Crew": "Affiliates",
    "Bearhawk - Aaron-Fee": "Affiliates",
    "Bearhawk - Ari-Fee": "Affiliates",
    "Bearhawk - Chris-Fee": "Affiliates",
    "Bearhawk - Dylan-Fee": "Affiliates",
    "Bearhawk - Gray-Fee": "Affiliates",
    "Bearhawk - Jason-Fee": "Affiliates",
    "Bearhawk Group-Fee": "Affiliates",
    "Best Tix-Fee": "Affiliates",
    "Damon and Crew-Fee": "Affiliates",
    "GK LLC-Fee": "Affiliates",
    "Jacks YS-Fee": "Affiliates",
    "Levovitz-Fee": "Affiliates",
    "Needle Tickets LLC-Fee": "Affiliates",
    "Pollak Tickets-Fee": "Affiliates",
    "Ticketwonders LLC-Fee": "Affiliates",
    "Yoni Levine-Fee": "Affiliates",
    "YourTickets-Fee": "Affiliates",
    "YS Katz-Fee": "Affiliates",
    "YS TL-Fee": "Affiliates",
    "YSA 2-Fee": "Affiliates",
    "YSA-Fee": "Affiliates",
    "YSM Tickets-Fee": "Affiliates",
    "YSS Tickets-Fee": "Affiliates",
    "YSW-Fee": "Affiliates",
    "YSA 3-Fee": "Affiliates",
    "The Ticket Guy-Fee": "Other",
    "Other Fees": "Other",
    "Due from/to Ticket Vault": "Other",
    "Due from/to TickPick": "Other",
    "YS Tickets-Fee": "Y&S - Deposit",
    "YS Tickets Spec-Fee": "Y&S - Deposit",
    "YS-Seatgeek-Fee": "Y&S - Deposit",
    "YS-Seatgeek2-Fee": "Y&S - Deposit",
    "David Mansbach": "Other",
    "GRA Investments": "Other",
    "GRA Investments - Brian": "Other",
    "Indiana Promotions": "Other",
    "Isaac Knopf": "Other",
    "JL": "Other",
    "Stuart Levy": "Other",
    "TV Test Company": "Other",
    "Best Tix": "Affiliates",
    "Ticketwonders LLC": "Affiliates",
    "Bearhawk - Aaron": "Affiliates",
    "Bearhawk - Ari": "Affiliates",
    "Bearhawk - Chris": "Affiliates",
    "Bearhawk - Dylan": "Affiliates",
    "Bearhawk - Gray": "Affiliates",
    "Bearhawk - Jason": "Affiliates",
    "Bearhawk Group": "Affiliates",
    "Not Found": "Other",
    "Upside LLC": "Other",
    "StubHub Loan": "Y&S - StubHub",
    "The Ticket Guy": "Other",
    "YS Tickets": "Y&S - RecPmt",
    "YS-SeatGeek2": "Y&S - RecPmt",
    "YS-Seatgeek2": "Y&S - RecPmt",
    "YS-Seatgeek": "Y&S - RecPmt",
    "YS-SeatGeek": "Y&S - RecPmt",
    "YS Tickets Spec": "Y&S - RecPmt",
    "YourTickets": "Affiliates",
    "YSA": "Affiliates",
    "YSA 2": "Affiliates",
    "YSA 3": "Affiliates",
    "Jacks YS": "Affiliates",
    "YS Katz": "Affiliates",
    "Yoni Levine": "Affiliates",
    "Levovitz": "Affiliates",
    "Needle Tickets LLC": "Affiliates",
    "YS TL": "Affiliates",
    "GK LLC": "Affiliates",
    "YSM Tickets": "Affiliates",
    "Pollak Tickets": "Affiliates",
    "YSS Tickets": "Affiliates",
    "YSW": "Affiliates",
    "Cancellation Fees": "Y&S - Deposit",
    "Slash Financial Inc": "Affiliates",
    "Slash Financial Inc-Fee": "Affiliates",
    "Ticketwonders2 LLC": "Affiliates",
    "Ticketwonders2 LLC-Fee": "Affiliates",
}

# Case-insensitive lookup dict
COMPANY_MAPPING_LOWER = {k.lower(): v for k, v in COMPANY_MAPPING.items()}

YSA_VARIANTS = {"YSA", "YSA 2", "YSA 3"}

HEADER_FILL = PatternFill("solid", start_color="375623", end_color="375623")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(name="Arial", bold=True, size=10)
DATA_FONT = Font(name="Arial", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

DATA_COLS = [
    "Company", "Date", "Network", "Type", "Order#", "Amount",
    "Performer", "Venue", "EventDate", "Section", "Row", "Seat", "Qty", "Reason"
]

COL_WIDTHS = {
    "Company": 20, "Date": 12, "Network": 14, "Type": 18,
    "Order#": 14, "Amount": 12, "Performer": 25, "Venue": 30,
    "EventDate": 22, "Section": 10, "Row": 6, "Seat": 10, "Qty": 6, "Reason": 25,
}


def parse_filename(filename):
    """Extract network and remittance date from filenames like:
       YS_Stubhub_5-5-26.csv, YS Stubhub 5-5-26.csv,
       GoTickets_05-13-2026.csv, GoTickets 05-13-2026.csv
    """
    base = os.path.splitext(filename)[0]
    base = base.replace(" ", "_")
    # Strip trailing copy indicators like (1), (2) etc before parsing
    import re as _re2
    base = _re2.sub(r"\(\d+\)$", "", base).rstrip("_")
    parts = base.split("_")

    import re as _re
    date_pat = _re.compile(r'^\d{1,2}-\d{1,2}-\d{2,4}$')

    # Find the date part — search from the end for first part matching date pattern
    date_idx = None
    for i in range(len(parts) - 1, -1, -1):
        if date_pat.match(parts[i]):
            date_idx = i
            break

    if date_idx is None:
        raise ValueError(f"Could not find a date in filename: {filename}")

    date_str = parts[date_idx]
    d = date_str.split("-")
    year = int(d[2]) if len(d[2]) == 4 else 2000 + int(d[2])
    remit_date = datetime(year, int(d[0]), int(d[1]))

    # Network is everything between the prefix (first part) and the date
    if date_idx >= 2:
        network_raw = "_".join(p for p in parts[1:date_idx] if p)
        prefix = parts[0]  # YS, YS2, TV, etc
    else:
        network_raw = parts[0]
        prefix = ""

    # Normalize key for lookups (lowercase, strip parens/spaces)
    network_key = network_raw.lower().replace("(", "").replace(")", "").replace(" ", "")

    # Display name map for tab network column
    network_display_map = {
        "vividseats": "Vivid Seats",
        "vividseatscad": "Vivid Seats (CAD)",
        "ticketevolution": "Ticket Evolution",
        "ticketsnow": "TicketsNow",
        "ticketsnowcad": "TicketsNow (CAD)",
    }
    network_display = network_display_map.get(network_key, network_raw)

    # Bank deposit network — CAD variants strip the (CAD) suffix
    deposit_network_map = {
        "vividseatscad": "Vivid Seats",
        "ticketsnowcad": "TicketsNow",
    }
    deposit_network = deposit_network_map.get(network_key, network_display)

    # Bank account map
    bank_account_map = {
        "ticketevolution": "EvoPay Main",
        "ticketsnowcad": "Wise Chkg (CAD)",
    }
    bank_account = bank_account_map.get(network_key, "FFB Chkg")

    return network_display, remit_date, deposit_network, bank_account, prefix


def _parse_event_date(val):
    """Parse EventDate safely — returns mm/dd/yyyy string or blank if unparseable."""
    if not pd.notna(val) or str(val).strip() == "":
        return ""
    try:
        dt = pd.to_datetime(val, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%m/%d/%Y")
    except Exception:
        return ""


def _is_valid_date(val):
    """Returns True if val can be parsed as a real date."""
    if not pd.notna(val) or str(val).strip() == "":
        return False
    try:
        dt = pd.to_datetime(val, errors="coerce")
        return not pd.isna(dt)
    except Exception:
        return False


def _resolve_shifted_cols(r):
    """Handle single or double column shifts caused by long EventName spilling over.
    Single shift: EventDate is non-date, Section has the real date.
    Double shift: EventDate is non-date, Section is also non-date, Row has the real date."""
    def _s(val):
        v = val if pd.notna(val) else ""
        return str(v).strip() if isinstance(v, str) else v

    def _reason_val():
        # Reason is the user-maintained column S. It stays at that fixed position
        # even when EventName commas shift the event-data columns, so always read
        # it directly rather than blanking it out on shifted rows.
        v = r.get("Reason", "")
        return str(v).strip() if pd.notna(v) and str(v).strip() != "" else ""

    event_date_valid = _is_valid_date(r.get("EventDate", ""))
    section_valid = _is_valid_date(r.get("Section", ""))
    row_valid = _is_valid_date(r.get("Row", ""))

    if event_date_valid:
        # No shift — normal row
        return {
            "Venue":     _s(r.get("Venue", "")),
            "EventDate": _parse_event_date(r.get("EventDate", "")),
            "Section":   _s(r.get("Section", "")),
            "Row":       _s(r.get("Row", "")),
            "Seat":      _s(r.get("Seat", "")),
            "Qty":       r.get("Qty", ""),
            "Reason":    _reason_val(),
        }
    elif section_valid:
        # Single shift — EventName spilled into one extra column
        return {
            "Venue":     _s(r.get("EventDate", "")),
            "EventDate": _parse_event_date(r.get("Section", "")),
            "Section":   _s(r.get("Row", "")),
            "Row":       _s(r.get("Seat", "")),
            "Seat":      _s(r.get("Qty", "")),
            "Qty":       r.get("CancellationReason", ""),
            "Reason":    _reason_val(),
        }
    elif row_valid:
        # Double shift — EventName spilled into two extra columns
        return {
            "Venue":     _s(r.get("Section", "")),
            "EventDate": _parse_event_date(r.get("Row", "")),
            "Section":   _s(r.get("Seat", "")),
            "Row":       _s(r.get("Qty", "")),
            "Seat":      _s(r.get("CancellationReason", "")),
            "Qty":       r.get("InternalFulfillmentStatus", ""),
            "Reason":    _reason_val(),
        }
    else:
        # Can't determine shift — return blanks for date-dependent fields
        return {
            "Venue":     _s(r.get("Venue", "")),
            "EventDate": "",
            "Section":   _s(r.get("Section", "")),
            "Row":       _s(r.get("Row", "")),
            "Seat":      _s(r.get("Seat", "")),
            "Qty":       r.get("Qty", ""),
            "Reason":    _reason_val(),
        }


def build_row(r, remit_date_str, network, evopay_sale=None, evopay_cancel=None):
    company_raw = str(r["Company"]).strip() if pd.notna(r["Company"]) else ""
    is_fee = company_raw.endswith("-Fee")
    is_cancellation = company_raw == "Cancellation Fees"
    amount = r["Amount"]

    if is_fee or is_cancellation:
        type_val = "Cancellation Fees"
    elif amount >= 0:
        type_val = "Payment"
    else:
        type_val = "Recoup"

    company_out = re.sub(r"-Fee$", "", company_raw)
    if company_out in YSA_VARIANTS:
        company_out = "YSA"
    if company_out.startswith("Bearhawk"):
        company_out = "Bearhawk Group"

    # Anything not explicitly mapped to Y&S, Affiliates, or StubHub falls to "Other"
    # (previously these rows were tagged "Unknown" and silently dropped from every tab).
    _key = company_raw.lower()
    if _key in COMPANY_MAPPING_LOWER:
        category = COMPANY_MAPPING_LOWER[_key]
    elif is_fee:
        # A "-Fee" company that isn't explicitly mapped inherits its base company's
        # category (e.g. "GK LLC-Fee" follows "GK LLC") instead of dropping to "Other".
        category = COMPANY_MAPPING_LOWER.get(re.sub(r"-fee$", "", _key), "Other")
    else:
        category = "Other"

    order_key = str(r["Order#"]).strip()
    sale_date = evopay_sale.get(order_key) if evopay_sale else None
    cancel_date = evopay_cancel.get(order_key) if evopay_cancel else None
    if amount < 0:
        # Canceled order: use the cancellation (Debit transfer) date, not the original sale.
        row_date = cancel_date or sale_date or remit_date_str
    else:
        row_date = sale_date or cancel_date or remit_date_str

    return {
        "Company": company_out,
        "Date": row_date,
        "Network": network,
        "Type": type_val,
        "Order#": str(int(r["Order#"])) if pd.notna(r["Order#"]) and str(r["Order#"]).isdigit() else (str(r["Order#"]) if pd.notna(r["Order#"]) else ""),
        "Amount": amount,
        "Performer": (
            r["Venue"] if (not _is_valid_date(r.get("EventDate", "")) and not _is_valid_date(r.get("Section", "")) and _is_valid_date(r.get("Row", ""))) and pd.notna(r.get("Venue"))
            else (r["Opponent"] if not _is_valid_date(r.get("EventDate", "")) and pd.notna(r.get("Opponent"))
            else (r["Performer"] if pd.notna(r["Performer"]) else ""))
        ),
        **_resolve_shifted_cols(r),
        "_category": category,
    }


def write_header_row(ws, row_num, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER


def write_data_cell(ws, row, col, value, fmt=None, align=ALIGN_LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell


def style_data_tab(ws, df):
    for col_idx, col_name in enumerate(DATA_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = ALIGN_CENTER

    for row_idx, row in df.iterrows():
        for col_idx, col_name in enumerate(DATA_COLS, 1):
            val = row[col_name]
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
            cell.font = DATA_FONT
            if col_name == "Amount":
                cell.number_format = "#,##0.00"
                cell.alignment = ALIGN_CENTER
            elif col_name in ("Date", "Network", "Type", "Order#", "EventDate", "Section", "Row", "Seat", "Qty"):
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_LEFT

    for col_idx, col_name in enumerate(DATA_COLS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DATA_COLS))}1"


# Network short codes for Deposit # (QBO 21-char limit)
NETWORK_SHORT_CODES = {
    "stubhub": "SH",
    "vivid seats": "VS",
    "ticket evolution": "TEVO",
    "ticketsnow": "TNOW",
    "gotickets": "GOTIX",
    "seatgeek": "SG",
    "gametime": "GT",
    "ticketnetwork": "TND",
    "mercury": "MERC",
}


def build_deposit_number(network_display, prefix, remit_date):
    """Build the short deposit number.
    Format: <SHORT>_<CAD?>_<PREFIX>_<MM-DD-YY>
    e.g. VS_YS2_06-12-26 or VS_CAD_YS2_06-12-26
    """
    import re as _re
    date_str = remit_date.strftime("%m-%d-%y")

    # Detect CAD
    is_cad = "(cad)" in network_display.lower() or " cad" in network_display.lower()

    # Clean network name for lookup — strip (CAD), (C), etc
    net_clean = _re.sub(r"\s*\(.*?\)", "", network_display).strip().lower()
    short = NETWORK_SHORT_CODES.get(net_clean, "")
    if not short:
        short = _re.sub(r"[^A-Z0-9]", "", network_display.upper())[:4]

    parts = [short]
    if is_cad:
        parts.append("CAD")
    parts.append(prefix)
    parts.append(date_str)
    return "_".join(parts)


def _allocate_fx(cad_total, usd_received, affiliate_amts, ys_bucket_amt):
    """Pro-rata split the CAD->USD shortfall (cad_total - usd_received) across each
    affiliate line plus a single Y&S bucket (receive payment + cancellation fees +
    other Y&S money). Returns (affiliate_shares: {broker: amt}, ys_share: float),
    rounded to cents and summing EXACTLY to the shortfall (residual penny lands on
    the largest line)."""
    shortfall = round(cad_total - usd_received, 2)
    if cad_total <= 0:
        return {}, 0.0
    lines = [(b, a) for b, a in affiliate_amts.items()]      # (broker, amount)
    if round(ys_bucket_amt, 2) != 0:
        lines.append((None, round(ys_bucket_amt, 2)))        # None -> Foreign Exchange Conversion
    shares = {key: round(amt / cad_total * shortfall, 2) for key, amt in lines}
    resid = round(shortfall - sum(shares.values()), 2)
    if resid and lines:
        biggest = max(lines, key=lambda kv: abs(kv[1]))[0]
        shares[biggest] = round(shares[biggest] + resid, 2)
    aff_shares = {k: v for k, v in shares.items() if k is not None}
    ys_share = shares.get(None, 0.0)
    return aff_shares, ys_share


# Leaf QBO account names for the CAD->USD conversion journal entry.
FX_USD_ACCOUNT = "Wise Chkg (USD)"
FX_CAD_ACCOUNT = "Wise Chkg (CAD)"
FX_CONVERSION_ACCOUNT = "Foreign Exchange Conversion"


def process(csv_path, filename, evopay_path=None, raw_df=None, usd_received=None):
    if raw_df is not None:
        raw = raw_df.copy()                # new two-zone path: rows already prepared
    else:
        raw = _read_csv_resilient(csv_path, usecols=range(19), engine="python", on_bad_lines="skip")
    raw.columns = raw.columns.str.strip()  # remove leading/trailing spaces from column names
    # Validate that column S (Reason) is present
    if "Reason" not in raw.columns:
        raise ValueError("Column S (Reason) is missing from this file. Please add the Reason column before uploading.")
    # Normalize Amount to numeric — some networks export it with thousands commas
    # (e.g. "-122,400.00"), which pandas would otherwise leave as text.
    if "Amount" in raw.columns:
        raw["Amount"] = raw["Amount"].apply(_to_amount)
    network_display, remit_date, deposit_network, bank_account, prefix = parse_filename(filename)
    remit_date_str = remit_date.strftime("%m/%d/%Y")
    network = network_display  # no (C) on detail tabs
    deposit_network_full = f"{deposit_network} (C)"  # (C) only on bank deposit
    memo = os.path.splitext(filename)[0]
    # Short deposit number for QBO and output files
    short_dep_num = build_deposit_number(network_display, prefix, remit_date)

    # Build EvoPay order->date lookups if provided.
    # An order can have TWO transfer rows: the original sale (a Credit) and,
    # if it was later canceled, the cancellation (a Debit). We keep them apart so
    # a negative (canceled) TE row can use the cancellation date, not the sale date.
    def _money(v):
        try:
            s = str(v).replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
            return float(s) if s else 0.0
        except Exception:
            return 0.0

    evopay_sale = {}     # order -> latest Credit-transfer date (original sale)
    evopay_cancel = {}   # order -> latest Debit-transfer date (cancellation)
    if evopay_path:
        try:
            if evopay_path.endswith('.csv'):
                ep = _read_csv_resilient(evopay_path)
            else:
                ep = pd.read_excel(evopay_path)
            ep.columns = ep.columns.str.strip()
            ep_transfers = ep[ep['Type'].astype(str).str.strip().str.lower() == 'transfer'].copy()
            _sale_dt, _cancel_dt = {}, {}
            for _, row in ep_transfers.iterrows():
                order = str(row['Order - PO #']).strip()
                dt = pd.to_datetime(str(row['Date Created']), errors='coerce')
                if pd.isna(dt):
                    continue
                debit = _money(row.get('Debit'))
                credit = _money(row.get('Credit'))
                if debit > 0:  # money out of the account = cancellation / claw-back
                    if order not in _cancel_dt or dt > _cancel_dt[order]:
                        _cancel_dt[order] = dt
                        evopay_cancel[order] = dt.strftime('%m/%d/%Y')
                else:          # money in (Credit) = original sale
                    if order not in _sale_dt or dt > _sale_dt[order]:
                        _sale_dt[order] = dt
                        evopay_sale[order] = dt.strftime('%m/%d/%Y')
        except Exception as e:
            print(f"Warning: could not read EvoPay file: {e}")

    rows = [build_row(r, remit_date_str, network, evopay_sale, evopay_cancel) for _, r in raw.iterrows()]
    df_out = pd.DataFrame(rows)

    # Drop $0.00 Cancellation Fees lines outright. They carry no money and just
    # clutter the detail tabs/sheets — TicketsNow reports a recoup line plus a fee
    # line per cancel and the fee is often $0, and the Problem Order "Already
    # Paid = Yes" split (line2 = amount + payout) is exactly $0 when the chargeback
    # is a pure payout recoup. Removing them doesn't change any deposit/receive
    # total (a $0 line sums to nothing).
    if not df_out.empty and "Type" in df_out.columns:
        _amt = pd.to_numeric(df_out["Amount"], errors="coerce").round(2)
        df_out = df_out[~((df_out["Type"] == "Cancellation Fees") & (_amt == 0))].reset_index(drop=True)

    ys_df = df_out[df_out["_category"].isin(["Y&S - Deposit", "Y&S - RecPmt"])].copy()
    affiliates_df = df_out[df_out["_category"] == "Affiliates"].copy()
    stubhub_df = df_out[df_out["_category"] == "Y&S - StubHub"].copy()
    other_df = df_out[df_out["_category"] == "Other"].copy()
    # TradeDesk Fees ("Other Fees") and Due from/to TickPick post to the bank
    # deposit but should NOT appear in the Other detail tab or the Other Google
    # Sheet — keep a detail-only view that drops them.
    OTHER_DEPOSIT_ONLY = {"Other Fees", "Due from/to TickPick"}
    other_df_detail = other_df[~other_df["Company"].isin(OTHER_DEPOSIT_ONLY)].copy()

    # ── Entry #1: Receive Payment ─────────────────────────────────────────────
    ys_payments = ys_df[ys_df["Type"] == "Payment"]["Amount"].sum()
    ys_recoups = ys_df[ys_df["Type"] == "Recoup"]["Amount"].sum()
    receive_payment_amt = round(ys_payments + ys_recoups, 2)

    # ── Entry #2: Bank Deposit ────────────────────────────────────────────────
    aff_grouped = affiliates_df.groupby("Company")["Amount"].sum().reset_index()
    aff_grouped.columns = ["Account", "Amount"]
    sh_grouped = stubhub_df.groupby("Company")["Amount"].sum().reset_index()
    sh_grouped.columns = ["Account", "Amount"]
    other_grouped = other_df.groupby("Company")["Amount"].sum().reset_index()
    other_grouped.columns = ["Account", "Amount"]
    ys_cancel_amt = round(ys_df[ys_df["Type"] == "Cancellation Fees"]["Amount"].sum(), 2)
    # Only emit a Cancellation Fees line when there actually are fees — otherwise a
    # $0.00 line shows up on the deposit (and would post a $0 line to QBO). Mirrors
    # the per-date (TE) path below.
    cancel_row = (pd.DataFrame([{"Account": "Cancellation Fees", "Amount": ys_cancel_amt}])
                  if ys_cancel_amt != 0 else pd.DataFrame(columns=["Account", "Amount"]))

    deposit_rows = pd.concat([aff_grouped, sh_grouped, other_grouped, cancel_row], ignore_index=True)
    deposit_rows["Amount"] = deposit_rows["Amount"].round(2)
    deposit_rows["Network"] = deposit_network_full
    deposit_rows["Date"] = remit_date_str
    deposit_rows["Deposit #"] = short_dep_num
    deposit_rows["Bank Account"] = bank_account

    bank_deposit_total = round(deposit_rows["Amount"].sum(), 2)
    combined_total = round(receive_payment_amt + bank_deposit_total, 2)

    # ── CAD->USD conversion (TicketsNow (CAD) etc.) ──────────────────────────
    # When the file is a CAD feed and the user supplies the USD actually received
    # from the bank conversion, allocate the shortfall pro rata: each affiliate
    # bears its own share (debited to its account); the Y&S bucket (receive
    # payment + cancellation fees + other Y&S money) goes to Foreign Exchange
    # Conversion. Produces a balanced QBO journal entry + per-affiliate FX rows
    # for the Affiliates sheet.
    is_cad = "(cad)" in network_display.lower() or " cad" in network_display.lower()
    fx_journal = None
    fx_detail_rows = []
    if is_cad and usd_received is not None:
        try:
            usd_amt = round(float(usd_received), 2)
        except (TypeError, ValueError):
            usd_amt = None
        cad_total = float(combined_total)
        if usd_amt is not None and cad_total > 0:
            affiliate_amts = {str(r["Account"]): round(float(r["Amount"]), 2)
                              for _, r in aff_grouped.iterrows() if round(float(r["Amount"]), 2) != 0}
            ys_bucket = float(round(receive_payment_amt
                              + (sh_grouped["Amount"].sum() if len(sh_grouped) else 0.0)
                              + (other_grouped["Amount"].sum() if len(other_grouped) else 0.0)
                              + ys_cancel_amt, 2))
            aff_shares, ys_share = _allocate_fx(cad_total, usd_amt, affiliate_amts, ys_bucket)

            # QBO journal entry lines (debits then the single CAD credit).
            # A positive share is borne as a debit; a negative share (a broker or
            # the Y&S bucket whose net payment was negative) must post as a POSITIVE
            # CREDIT, never a negative debit. This keeps the entry balanced and the
            # signs correct on the books.
            def _fx_line(account, amt):
                amt = round(amt, 2)
                if amt >= 0:
                    return {"account": account, "debit": float(amt)}
                return {"account": account, "credit": float(-amt)}

            jlines = [{"account": FX_USD_ACCOUNT, "debit": float(usd_amt)}]
            for broker, share in aff_shares.items():
                if round(share, 2):
                    jlines.append(_fx_line(broker, share))
            if round(ys_share, 2):
                jlines.append(_fx_line(FX_CONVERSION_ACCOUNT, ys_share))
            jlines.append({"account": FX_CAD_ACCOUNT, "credit": float(round(cad_total, 2))})
            fx_journal = {
                "date": remit_date_str,
                "deposit_num": short_dep_num,
                "network": network_display,
                "usd_received": float(usd_amt),
                "cad_total": float(round(cad_total, 2)),
                "shortfall": float(round(cad_total - usd_amt, 2)),
                "lines": jlines,
            }
            # Per-affiliate FX rows for the Affiliates Google Sheet (negative share).
            for broker, share in aff_shares.items():
                if round(share, 2):
                    fx_detail_rows.append({
                        "Tab": "Affiliates",
                        "Company": broker, "Date": remit_date_str,
                        "Network": network_display, "Type": "FX",
                        "Order#": "", "Amount": float(round(-share, 2)),
                        "Performer": "", "Venue": "", "EventDate": "",
                        "Section": "", "Row": "", "Seat": "", "Qty": "", "Reason": "",
                    })

    # ── Determine if this is a TE (per-date) file ────────────────────────────
    is_te = bool(evopay_sale or evopay_cancel)

    # ── Build per-date summary rows for TE files ──────────────────────────────
    def _get_filename_prefix(memo_str):
        """Extract everything before the date portion of the filename, stripping trailing underscore."""
        import re as _re
        m = _re.search(r'_\d{1,2}-\d{1,2}-\d{2,4}', memo_str)
        return memo_str[:m.start()].rstrip('_') if m else memo_str

    def _date_to_filename_fmt(date_str):
        """Convert mm/dd/yyyy to MM-DD-YY for use in filenames."""
        try:
            dt = pd.to_datetime(date_str)
            return dt.strftime("%m-%d-%y")
        except Exception:
            return date_str

    if is_te:
        fn_prefix = _get_filename_prefix(memo)
        # Dates actually assigned to rows (sale dates for sales, cancellation
        # dates for canceled orders) — this keeps the range/filename in-range.
        all_dates = sorted(set(df_out["Date"].dropna().astype(str)))

        # Per-date Receive Payment rows (exclude dates with zero amount)
        rp_rows = []
        for d in all_dates:
            d_rows = ys_df[ys_df["Date"] == d]
            pay = d_rows[d_rows["Type"] == "Payment"]["Amount"].sum()
            rec = d_rows[d_rows["Type"] == "Recoup"]["Amount"].sum()
            amt = round(pay + rec, 2)
            if amt == 0:
                continue
            d_date = pd.to_datetime(d)
            dep_num = build_deposit_number(network_display, prefix, d_date)
            rp_rows.append({"Date": d, "Amount": amt, "Deposit #": dep_num})

        # Per-date Bank Deposit rows
        bd_rows_by_date = []
        for d in all_dates:
            d_date = pd.to_datetime(d)
            dep_num = build_deposit_number(network_display, prefix, d_date)
            # Affiliates
            aff_d = affiliates_df[affiliates_df["Date"] == d].groupby("Company")["Amount"].sum().reset_index()
            aff_d.columns = ["Account", "Amount"]
            # StubHub loan
            sh_d = stubhub_df[stubhub_df["Date"] == d].groupby("Company")["Amount"].sum().reset_index()
            sh_d.columns = ["Account", "Amount"]
            # Other
            oth_d = other_df[other_df["Date"] == d].groupby("Company")["Amount"].sum().reset_index()
            oth_d.columns = ["Account", "Amount"]
            # Cancellation fees
            cancel_d = round(ys_df[(ys_df["Date"] == d) & (ys_df["Type"] == "Cancellation Fees")]["Amount"].sum(), 2)
            cancel_row_d = pd.DataFrame([{"Account": "Cancellation Fees", "Amount": cancel_d}]) if cancel_d != 0 else pd.DataFrame(columns=["Account","Amount"])
            d_rows = pd.concat([aff_d, sh_d, oth_d, cancel_row_d], ignore_index=True)
            d_rows["Amount"] = d_rows["Amount"].round(2)
            d_rows["Network"] = deposit_network_full
            d_rows["Date"] = d
            d_rows["Deposit #"] = dep_num
            d_rows["Bank Account"] = bank_account
            bd_rows_by_date.append(d_rows)

        all_bd_rows = pd.concat(bd_rows_by_date, ignore_index=True)

        # Date range for filenames — use only dates with actual activity (non-zero RP or BD rows)
        active_dates = set()
        for rp in rp_rows:
            active_dates.add(rp["Date"])
        for _, row in all_bd_rows.iterrows():
            if row["Amount"] != 0:
                active_dates.add(row["Date"])
        active_dates_sorted = sorted(active_dates)
        if active_dates_sorted:
            min_date = _date_to_filename_fmt(active_dates_sorted[0])
            max_date = _date_to_filename_fmt(active_dates_sorted[-1])
        else:
            min_date = _date_to_filename_fmt(all_dates[0])
            max_date = _date_to_filename_fmt(all_dates[-1])
        date_range_str = f"{min_date} to {max_date}"
    else:
        date_range_str = None

    # ── Build Applied Payments workbook ───────────────────────────────────────
    wb1 = openpyxl.Workbook()
    wb1.remove(wb1.active)

    # Summary tab
    ws_sum = wb1.create_sheet("Summary")
    SUM_COLS = ["Memo", "Amount", "Network", "Date", "Deposit #", "Bank Account"]
    BD_COLS  = ["Account", "Amount", "Network", "Date", "Deposit #", "Bank Account"]
    FX_COLS  = ["Account", "Amount", "Network", "Date", "Deposit #"]
    SUM_WIDTHS = [28, 14, 20, 12, 32, 14]

    if is_te:
        ws_sum.cell(row=1, column=1, value="Receive Payment").font = SECTION_FONT
        write_header_row(ws_sum, 2, SUM_COLS)
        cur_row = 3
        for rp in rp_rows:
            write_data_cell(ws_sum, cur_row, 1, rp["Deposit #"])
            write_data_cell(ws_sum, cur_row, 2, rp["Amount"], fmt="#,##0.00", align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 3, deposit_network_full, align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 4, rp["Date"], align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 5, rp["Deposit #"])
            write_data_cell(ws_sum, cur_row, 6, bank_account, align=ALIGN_CENTER)
            cur_row += 1
        cur_row += 1  # blank row
        ws_sum.cell(row=cur_row, column=1, value="Bank Deposit").font = SECTION_FONT
        cur_row += 1
        write_header_row(ws_sum, cur_row, BD_COLS)
        cur_row += 1
        for _, row in all_bd_rows.iterrows():
            write_data_cell(ws_sum, cur_row, 1, row["Account"])
            write_data_cell(ws_sum, cur_row, 2, row["Amount"], fmt="#,##0.00", align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 3, row["Network"], align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 4, row["Date"], align=ALIGN_CENTER)
            write_data_cell(ws_sum, cur_row, 5, row["Deposit #"])
            write_data_cell(ws_sum, cur_row, 6, row["Bank Account"], align=ALIGN_CENTER)
            cur_row += 1
    else:
        ws_sum.cell(row=1, column=1, value="Receive Payment").font = SECTION_FONT
        write_header_row(ws_sum, 2, SUM_COLS)
        # Header always present; the data row only when there's an actual payment.
        if receive_payment_amt != 0:
            write_data_cell(ws_sum, 3, 1, short_dep_num)
            write_data_cell(ws_sum, 3, 2, receive_payment_amt, fmt="#,##0.00", align=ALIGN_CENTER)
            write_data_cell(ws_sum, 3, 3, deposit_network_full, align=ALIGN_CENTER)
            write_data_cell(ws_sum, 3, 4, remit_date_str, align=ALIGN_CENTER)
            write_data_cell(ws_sum, 3, 5, short_dep_num)
            write_data_cell(ws_sum, 3, 6, bank_account, align=ALIGN_CENTER)
        ws_sum.cell(row=5, column=1, value="Bank Deposit").font = SECTION_FONT
        write_header_row(ws_sum, 6, BD_COLS)
        for i, row in deposit_rows.iterrows():
            r = 7 + i
            write_data_cell(ws_sum, r, 1, row["Account"])
            write_data_cell(ws_sum, r, 2, row["Amount"], fmt="#,##0.00", align=ALIGN_CENTER)
            write_data_cell(ws_sum, r, 3, row["Network"], align=ALIGN_CENTER)
            write_data_cell(ws_sum, r, 4, row["Date"], align=ALIGN_CENTER)
            write_data_cell(ws_sum, r, 5, row["Deposit #"])
            write_data_cell(ws_sum, r, 6, row["Bank Account"], align=ALIGN_CENTER)
        # FX Journal Entry section (TicketsNow CAD only). Positive = debit,
        # negative = credit; the column total nets to zero (balanced entry).
        if fx_journal:
            fx_row = 8 + len(deposit_rows)        # one blank row after the bank deposit
            ws_sum.cell(row=fx_row, column=1, value="FX Journal Entry").font = SECTION_FONT
            fx_row += 1
            write_header_row(ws_sum, fx_row, FX_COLS)
            fx_row += 1
            for line in fx_journal["lines"]:
                amt = line["debit"] if "debit" in line else -line["credit"]
                write_data_cell(ws_sum, fx_row, 1, line["account"])
                write_data_cell(ws_sum, fx_row, 2, amt, fmt="#,##0.00", align=ALIGN_CENTER)
                write_data_cell(ws_sum, fx_row, 3, fx_journal["network"], align=ALIGN_CENTER)
                write_data_cell(ws_sum, fx_row, 4, fx_journal["date"], align=ALIGN_CENTER)
                write_data_cell(ws_sum, fx_row, 5, fx_journal["deposit_num"])
                fx_row += 1
    for col_idx, w in enumerate(SUM_WIDTHS, 1):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = w

    # Data tabs. Y&S and Affiliates are always present; StubHub Loan and Other
    # appear only when they actually have rows.
    tab_data = {
        "Y&S": ys_df[DATA_COLS].reset_index(drop=True),
        "Affiliates": affiliates_df[DATA_COLS].reset_index(drop=True),
        "StubHub Loan": stubhub_df[DATA_COLS].reset_index(drop=True),
        "Other": other_df_detail[DATA_COLS].reset_index(drop=True),
    }
    optional_tabs = {"StubHub Loan", "Other"}
    for tab_name, df in tab_data.items():
        if tab_name in optional_tabs and df.empty:
            continue
        ws = wb1.create_sheet(tab_name)
        style_data_tab(ws, df)

    # ── Build Bank Deposit workbook ───────────────────────────────────────────
    bd_cols = ["Account", "Amount", "Network", "Date", "Deposit #", "Bank Account"]
    bd_col_widths = [28, 14, 20, 12, 32, 14]
    bd_source = all_bd_rows if is_te else deposit_rows

    wb2 = openpyxl.Workbook()
    ws_bd = wb2.active
    ws_bd.title = "Bank Deposit"
    write_header_row(ws_bd, 1, bd_cols)
    for i, row in bd_source.iterrows():
        r = 2 + i
        write_data_cell(ws_bd, r, 1, row["Account"])
        write_data_cell(ws_bd, r, 2, row["Amount"], fmt="#,##0.00", align=ALIGN_CENTER)
        write_data_cell(ws_bd, r, 3, row["Network"], align=ALIGN_CENTER)
        write_data_cell(ws_bd, r, 4, row["Date"], align=ALIGN_CENTER)
        write_data_cell(ws_bd, r, 5, row["Deposit #"])
        write_data_cell(ws_bd, r, 6, row["Bank Account"], align=ALIGN_CENTER)
    for col_idx, w in enumerate(bd_col_widths, 1):
        ws_bd.column_dimensions[get_column_letter(col_idx)].width = w
    ws_bd.freeze_panes = "A2"
    ws_bd.auto_filter.ref = f"A1:{get_column_letter(len(bd_cols))}1"

    # ── Build Receive Payment workbook ────────────────────────────────────────
    # Always produced with its header row so the layout is consistent; a data row
    # appears only when there's an actual receive payment. A $0 receive payment
    # yields a headers-only sheet (and nothing is pushed to QBO).
    if is_te:
        rp_rows_wb = rp_rows                      # per-date list, already excludes $0 dates
    else:
        rp_rows_wb = ([{"Date": remit_date_str, "Amount": receive_payment_amt,
                        "Deposit #": short_dep_num}]
                      if receive_payment_amt != 0 else [])

    rp_cols = ["Memo", "Amount", "Network", "Date", "Deposit #", "Bank Account"]
    rp_col_widths = [32, 14, 20, 12, 32, 14]
    wb3 = openpyxl.Workbook()
    ws_rp = wb3.active
    ws_rp.title = "Receive Payment"
    write_header_row(ws_rp, 1, rp_cols)
    for i, rp in enumerate(rp_rows_wb):
        r = 2 + i
        write_data_cell(ws_rp, r, 1, rp["Deposit #"])
        write_data_cell(ws_rp, r, 2, rp["Amount"], fmt="#,##0.00", align=ALIGN_CENTER)
        write_data_cell(ws_rp, r, 3, deposit_network_full, align=ALIGN_CENTER)
        write_data_cell(ws_rp, r, 4, rp["Date"], align=ALIGN_CENTER)
        write_data_cell(ws_rp, r, 5, rp["Deposit #"])
        write_data_cell(ws_rp, r, 6, bank_account, align=ALIGN_CENTER)
    for col_idx, w in enumerate(rp_col_widths, 1):
        ws_rp.column_dimensions[get_column_letter(col_idx)].width = w
    ws_rp.freeze_panes = "A2"
    ws_rp.auto_filter.ref = f"A1:{get_column_letter(len(rp_cols))}1"

    # ── Detail rows for Google Sheets (the four data tabs, combined) ──────────
    detail_source = [
        ("Y&S", ys_df),
        ("Affiliates", affiliates_df),
        ("StubHub Loan", stubhub_df),
        ("Other", other_df_detail),
    ]
    detail_rows_data = []
    for tab_name, df in detail_source:
        for _, r in df[DATA_COLS].iterrows():
            rec = {"Tab": tab_name}
            for col in DATA_COLS:
                v = r[col]
                if pd.isna(v):
                    v = ""
                elif hasattr(v, "item"):   # numpy scalar -> native python
                    v = v.item()
                rec[col] = v
            detail_rows_data.append(rec)

    # Append the per-affiliate FX rows (CAD files only) so they ride along to the
    # Affiliates sheet on the normal Google Sheets push.
    detail_rows_data.extend(fx_detail_rows)

    return {
        "wb_applied": wb1,
        "wb_deposit": wb2,
        "wb_receive": wb3,
        "memo": memo,
        "date_range_str": date_range_str,
        "receive_payment_amt": receive_payment_amt,
        "bank_deposit_total": bank_deposit_total,
        "combined_total": combined_total,
        "bank_account": bank_account,
        "deposit_network_full": deposit_network_full,
        "network_display": network_display,
        "is_cad": is_cad,
        "fx_journal": fx_journal,
        # Raw data for QBO push
        "all_bd_rows_data": (all_bd_rows if is_te else deposit_rows).to_dict("records"),
        "rp_rows_data": rp_rows_wb,
        # Detail-tab rows for the Google Sheet append
        "detail_rows_data": detail_rows_data,
    }
